#!/usr/bin/env python3
"""
license-prep-v2.py — 按 missions_cursorlicense_prep_v2.md 要求处理 data/input.xlsx：

1. 从 data/input.xlsx 读取
2. 从 Cursor_Credits_Justification 标签页提取 Note 非空的 users_to_add，输出为 justified.csv
3. 匹配 Token Usage 的 Email 与 Cursor Licenses 的 users_to_add（不区分大小写），
   追加 Total Prompts、Fast Premium Prompts、On-Demand Spend 三列
4. 新增 Total 列 = Total Prompts + Fast Premium Prompts，按 Total 升序排序
5. 两轮循环计算 monthly-spend-limit：
   第一轮（仅对 justified.csv 中的用户赋值）：
     On-Demand Spend > 100 → 150
     On-Demand Spend > 50  → 100
     On-Demand Spend > 20  → 50
     On-Demand Spend > 10  → 30
     On-Demand Spend > 2   → 15
     否则 → 不变 (0)
   第二轮（仅对 justified.csv 中的用户叠加）：
     Total < 400  → 不变
     Total < 600  → +10
     Total < 1000 → +40
     Total < 1500 → +80
     Total < 2000 → +120
     Total >= 2000 → 直接设为 200
6. 结果保存到新表「Cursor Licenses New」

Environment: Python 3.12，与项目其他 .py 共用同一虚拟环境。
"""

import csv
from pathlib import Path

import openpyxl

INPUT_PATH = Path(__file__).resolve().parent / "data" / "input.xlsx"
JUSTIFIED_CSV = Path(__file__).resolve().parent / "data" / "justified.csv"


def find_col(header_row, name_candidates):
    """在首行中查找列名（不区分大小写），支持多个候选名，返回第一个匹配的列号。"""
    for cell in header_row:
        val = (cell.value or "").strip()
        for candidate in name_candidates:
            if val.lower() == candidate.lower():
                return cell.column
    return None


def find_col_nth(header_row, name, nth=1):
    """查找第 nth 个匹配的列名，用于处理重复列名。"""
    count = 0
    for cell in header_row:
        val = (cell.value or "").strip()
        if val.lower() == name.lower():
            count += 1
            if count == nth:
                return cell.column
    return None


def to_number(val, default=0):
    """将单元格值转为数字，失败时返回 default。"""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def main():
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"输入文件不存在: {INPUT_PATH}")

    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)

    # -------------------------------------------------------------------------
    # 步骤 2: 从 Cursor_Credits_Justification 提取 Note 非空的邮箱，输出 justified.csv
    # -------------------------------------------------------------------------
    if "Cursor_Credits_Justification" not in wb.sheetnames:
        raise ValueError("工作簿中未找到工作表 'Cursor_Credits_Justification'")
    ws_just = wb["Cursor_Credits_Justification"]
    header_just = list(next(ws_just.iter_rows(min_row=1, max_row=1)))

    col_users_j = find_col(header_just, ["users_to_add"])
    col_note_j  = find_col(header_just, ["Note"])
    if col_users_j is None:
        raise ValueError("'Cursor_Credits_Justification' 中未找到 'users_to_add' 列")
    if col_note_j is None:
        raise ValueError("'Cursor_Credits_Justification' 中未找到 'Note' 列")

    # 提取 Note 非空的邮箱（小写集合，用于后续匹配）
    justified_emails = set()
    justified_rows = []
    for row in ws_just.iter_rows(min_row=2, values_only=True):
        email = (row[col_users_j - 1] or "").strip()
        note  = row[col_note_j - 1]
        if email and note is not None and str(note).strip():
            justified_emails.add(email.lower())
            justified_rows.append(email)

    # 写出 justified.csv
    JUSTIFIED_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(JUSTIFIED_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["users_to_add"])
        for email in justified_rows:
            writer.writerow([email])
    print(f"justified.csv 已写出，共 {len(justified_rows)} 条记录: {JUSTIFIED_CSV}")

    # -------------------------------------------------------------------------
    # 步骤 3: 从 Token Usage 读取 Email → (Total Prompts, Fast Premium Prompts, On-Demand Spend)
    # Token Usage 中存在两列同名 On-Demand Spend，取第 2 列（列8）作为实际花费
    # -------------------------------------------------------------------------
    if "Token Usage" not in wb.sheetnames:
        raise ValueError("工作簿中未找到工作表 'Token Usage'")
    ws_token = wb["Token Usage"]
    header_token = list(next(ws_token.iter_rows(min_row=1, max_row=1)))

    col_email_t = find_col(header_token, ["Email"])
    col_total_t = find_col(header_token, ["Total Prompts", "Total Prompots"])
    col_fast_t  = find_col(header_token, ["Fast Premium Prompts"])
    # Token Usage 中有两列 On-Demand Spend，取第 2 列（实际花费）
    col_spend_t = find_col_nth(header_token, "On-Demand Spend", nth=2)
    if col_spend_t is None:
        # 若只有一列，则取第一列
        col_spend_t = find_col(header_token, ["On-Demand Spend"])

    if col_email_t is None:
        raise ValueError("'Token Usage' 中未找到 'Email' 列")
    if col_total_t is None:
        raise ValueError("'Token Usage' 中未找到 'Total Prompts' 列")
    if col_fast_t is None:
        raise ValueError("'Token Usage' 中未找到 'Fast Premium Prompts' 列")
    if col_spend_t is None:
        raise ValueError("'Token Usage' 中未找到 'On-Demand Spend' 列")

    email_data = {}  # email_lower → {"total": int, "fast": int, "spend": float}
    for row in ws_token.iter_rows(min_row=2):
        email = (row[col_email_t - 1].value or "").strip()
        if not email:
            continue
        email_data[email.lower()] = {
            "total": int(to_number(row[col_total_t - 1].value)),
            "fast":  int(to_number(row[col_fast_t  - 1].value)),
            "spend": to_number(row[col_spend_t - 1].value),
        }

    # -------------------------------------------------------------------------
    # 步骤 3 续 + 步骤 4: 读取 Cursor Licenses，匹配并追加三列，新增 Total，升序排序
    # -------------------------------------------------------------------------
    if "Cursor Licenses" not in wb.sheetnames:
        raise ValueError("工作簿中未找到工作表 'Cursor Licenses'")
    ws_lic  = wb["Cursor Licenses"]
    max_col = ws_lic.max_column
    header_lic = list(next(ws_lic.iter_rows(min_row=1, max_row=1, max_col=max_col)))

    col_users = find_col(header_lic, ["users_to_add", "Users_to_add"])
    col_limit = find_col(header_lic, ["monthly-spend-limit"])
    if col_users is None:
        raise ValueError("'Cursor Licenses' 中未找到 'users_to_add' 列")
    if col_limit is None:
        raise ValueError("'Cursor Licenses' 中未找到 'monthly-spend-limit' 列")

    # 读入数据行，跳过 users_to_add 为空的行
    # 每条记录: (row_vals, total_prompts, fast_premium, spend)
    data_rows = []
    for r in range(2, ws_lic.max_row + 1):
        row_vals = [ws_lic.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        user = (row_vals[col_users - 1] or "").strip()
        if not user:
            continue
        info  = email_data.get(user.lower(), {})
        total = info.get("total", 0)
        fast  = info.get("fast", 0)
        spend = info.get("spend", 0)
        data_rows.append((row_vals, total, fast, spend))

    # 按 Total (= Total Prompts + Fast Premium Prompts) 升序排序
    data_rows.sort(key=lambda x: (x[1] + x[2], (x[0][col_users - 1] or "")))

    # -------------------------------------------------------------------------
    # 步骤 5: 两轮循环计算 monthly-spend-limit
    # -------------------------------------------------------------------------
    limits = []
    for row_vals, total_prompts, fast_premium, spend in data_rows:
        combined = total_prompts + fast_premium
        user_email = (row_vals[col_users - 1] or "").strip().lower()

        # 第一轮：仅对 justified.csv 中的用户按 On-Demand Spend 设底值
        if user_email in justified_emails:
            if spend > 100:
                base = 150
            elif spend > 50:
                base = 100
            elif spend > 20:
                base = 50
            elif spend > 10:
                base = 30
            elif spend > 2:
                base = 15
            else:
                base = 0
        else:
            base = 0

        # 第二轮：仅对 justified.csv 中的用户按 Total 叠加（>= 2000 直接设为 200）
        if user_email not in justified_emails:
            final = base  # 不在 justified.csv 中，跳过第二轮
        elif combined >= 2000:
            final = 200
        elif combined >= 1500:
            final = base + 120
        elif combined >= 1000:
            final = base + 80
        elif combined >= 600:
            final = base + 40
        elif combined >= 400:
            final = base + 10
        else:
            final = base  # Total < 400 不变

        limits.append(final)

    # -------------------------------------------------------------------------
    # 步骤 6: 写出到新表「Cursor Licenses New」（不修改原表其他内容）
    # -------------------------------------------------------------------------
    wb2 = openpyxl.load_workbook(INPUT_PATH, data_only=False)
    if "Cursor Licenses New" in wb2.sheetnames:
        del wb2["Cursor Licenses New"]
    ws_new = wb2.create_sheet("Cursor Licenses New")

    # 表头：原表头 + Total Prompts + Fast Premium Prompts + On-Demand Spend + Total
    for c, cell in enumerate(header_lic, start=1):
        ws_new.cell(row=1, column=c, value=cell.value)
    col_total_new = max_col + 1
    col_fast_new  = max_col + 2
    col_spend_new = max_col + 3
    col_comb_new  = max_col + 4
    ws_new.cell(row=1, column=col_total_new, value="Total Prompts")
    ws_new.cell(row=1, column=col_fast_new,  value="Fast Premium Prompts")
    ws_new.cell(row=1, column=col_spend_new, value="On-Demand Spend")
    ws_new.cell(row=1, column=col_comb_new,  value="Total")

    for row_idx, ((row_vals, total_prompts, fast_premium, spend), limit_val) in enumerate(
        zip(data_rows, limits), start=2
    ):
        for c, val in enumerate(row_vals, start=1):
            ws_new.cell(row=row_idx, column=c, value=limit_val if c == col_limit else val)
        ws_new.cell(row=row_idx, column=col_total_new, value=total_prompts)
        ws_new.cell(row=row_idx, column=col_fast_new,  value=fast_premium)
        ws_new.cell(row=row_idx, column=col_spend_new, value=spend)
        ws_new.cell(row=row_idx, column=col_comb_new,  value=total_prompts + fast_premium)

    wb2.save(INPUT_PATH)
    print(f"结果已写入工作表「Cursor Licenses New」，共 {len(data_rows)} 行，文件: {INPUT_PATH}")

    # 输出摘要统计
    non_zero = sum(1 for l in limits if l > 0)
    print(f"monthly-spend-limit 非零记录数: {non_zero}/{len(limits)}")
    print(f"justified.csv 收录用户数: {len(justified_rows)}")


if __name__ == "__main__":
    main()
