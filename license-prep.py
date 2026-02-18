#!/usr/bin/env python3
"""
license-prep.py — 按 missions.md 要求处理 data/input.xlsx：

1. 从 data\\input.xlsx 读取
2. 用 Token Usage 的 Email 匹配 Cursor Licenses 的 Users_to_add（不区分大小写），
   追加 Total Prompts、Fast Premium Prompts、On-Demand Spend 三列
3. 新增 Total 列 = Total Prompts + Fast Premium Prompts，按 Total 升序排序
4. 先将所有 monthly-spend-limit 视为 0，再经两轮循环赋值：
   第一轮（On-Demand Spend 设底值）：>50→100, >20→80, >10→40, >2→20，否则 0
   第二轮（Total 叠加）：<300 不变, <500 +20, <1000 +60, <1500 +100, <2000 +150, >=2000 直接设为 200
5. 结果保存到新表「Cursor Licenses New」（不修改原表）

Environment: Python 3.12，与项目其他 .py 共用同一虚拟环境。
"""

from pathlib import Path
import openpyxl

INPUT_PATH = Path(__file__).resolve().parent / "data" / "input.xlsx"


def find_col(header_row, name_candidates):
    """在首行中查找列名（不区分大小写），支持多个候选名。"""
    for cell in header_row:
        val = (cell.value or "").strip()
        for candidate in name_candidates:
            if val.lower() == candidate.lower():
                return cell.column
    return None


def to_number(val, default=0):
    """将单元格值转为数字，失败时返回 default。"""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def main():
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"输入文件不存在: {INPUT_PATH}")

    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)

    # --- 1) 从 Token Usage 读取 Email -> (Total Prompts, Fast Premium Prompts, On-Demand Spend) ---
    if "Token Usage" not in wb.sheetnames:
        raise ValueError("工作簿中未找到工作表 'Token Usage'")
    ws_token = wb["Token Usage"]
    header_token = list(next(ws_token.iter_rows(min_row=1, max_row=1)))

    col_email_t   = find_col(header_token, ["Email"])
    col_total_t   = find_col(header_token, ["Total Prompts", "Total Prompots"])
    col_fast_t    = find_col(header_token, ["Fast Premium Prompts"])
    col_spend_t   = find_col(header_token, ["On-Demand Spend"])

    if col_email_t is None or col_total_t is None:
        raise ValueError("'Token Usage' 中未找到 'Email' 或 'Total Prompts' 列")
    if col_fast_t is None:
        raise ValueError("'Token Usage' 中未找到 'Fast Premium Prompts' 列")
    if col_spend_t is None:
        raise ValueError("'Token Usage' 中未找到 'On-Demand Spend' 列")

    email_data = {}   # email_lower -> {"total": int, "fast": int, "spend": float}
    for row in ws_token.iter_rows(min_row=2):
        email = (row[col_email_t - 1].value or "").strip()
        if not email:
            continue
        email_data[email.lower()] = {
            "total": int(to_number(row[col_total_t - 1].value)),
            "fast":  int(to_number(row[col_fast_t - 1].value)),
            "spend": to_number(row[col_spend_t - 1].value),
        }

    # --- 2) 读取 Cursor Licenses，匹配并追加三列 ---
    if "Cursor Licenses" not in wb.sheetnames:
        raise ValueError("工作簿中未找到工作表 'Cursor Licenses'")
    ws_lic  = wb["Cursor Licenses"]
    max_col = ws_lic.max_column
    header_lic = list(next(ws_lic.iter_rows(min_row=1, max_row=1, max_col=max_col)))

    col_users = find_col(header_lic, ["Users_to_add"])
    col_limit = find_col(header_lic, ["monthly-spend-limit"])
    if col_users is None:
        raise ValueError("'Cursor Licenses' 中未找到 'Users_to_add' 列")
    if col_limit is None:
        raise ValueError("'Cursor Licenses' 中未找到 'monthly-spend-limit' 列")

    # 读入数据行，跳过 Users_to_add 为空的行（含 Excel 追踪的空行）
    # 每条记录：(row_vals, total_prompts, fast_premium, spend)
    data_rows = []
    for r in range(2, ws_lic.max_row + 1):
        row_vals = [ws_lic.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        user = (row_vals[col_users - 1] or "").strip()
        if not user:
            continue
        info  = email_data.get(user.lower(), {})
        total = info.get("total", 0)
        fast  = info.get("fast", 0)
        spend = info.get("spend", 0)
        data_rows.append((row_vals, total, fast, spend))

    # --- 3) 新增 Total = Total Prompts + Fast Premium Prompts，按 Total 升序排序 ---
    data_rows.sort(key=lambda x: (x[1] + x[2], (x[0][col_users - 1] or "")))

    # --- 4) 两轮循环计算 monthly-spend-limit ---
    limits = []
    for _, total_prompts, fast_premium, spend in data_rows:
        combined = total_prompts + fast_premium

        # 第一轮：On-Demand Spend (>) 设底值
        if spend > 50:
            base = 100
        elif spend > 20:
            base = 80
        elif spend > 10:
            base = 40
        elif spend > 2:
            base = 20
        else:
            base = 0

        # 第二轮：Total (combined) 叠加（>=2000 直接覆盖为 200）
        if combined < 300:
            final = base           # 不变
        elif combined < 500:
            final = base + 20
        elif combined < 1000:
            final = base + 60
        elif combined < 1500:
            final = base + 100
        elif combined < 2000:
            final = base + 150
        else:
            final = 200            # 直接设为 200，不再叠加

        limits.append(final)

    # --- 5) 写出到新表「Cursor Licenses New」（不修改原表）---
    wb2 = openpyxl.load_workbook(INPUT_PATH, data_only=False)
    if "Cursor Licenses New" in wb2.sheetnames:
        del wb2["Cursor Licenses New"]
    ws_new = wb2.create_sheet("Cursor Licenses New")

    # 表头：原表头 + Total Prompts + Fast Premium Prompts + On-Demand Spend + Total
    for c, cell in enumerate(header_lic, start=1):
        ws_new.cell(row=1, column=c, value=cell.value)
    col_total_new = max_col + 1
    col_fast_new  = max_col + 2
    col_spend_new = max_col + 3
    col_comb_new  = max_col + 4
    ws_new.cell(row=1, column=col_total_new, value="Total Prompts")
    ws_new.cell(row=1, column=col_fast_new,  value="Fast Premium Prompts")
    ws_new.cell(row=1, column=col_spend_new, value="On-Demand Spend")
    ws_new.cell(row=1, column=col_comb_new,  value="Total")

    for row_idx, ((row_vals, total_prompts, fast_premium, spend), limit_val) in enumerate(
        zip(data_rows, limits), start=2
    ):
        for c, val in enumerate(row_vals, start=1):
            ws_new.cell(row=row_idx, column=c, value=limit_val if c == col_limit else val)
        ws_new.cell(row=row_idx, column=col_total_new, value=total_prompts)
        ws_new.cell(row=row_idx, column=col_fast_new,  value=fast_premium)
        ws_new.cell(row=row_idx, column=col_spend_new, value=spend)
        ws_new.cell(row=row_idx, column=col_comb_new,  value=total_prompts + fast_premium)

    wb2.save(INPUT_PATH)
    print("结果已写入工作表「Cursor Licenses New」，文件:", INPUT_PATH)


if __name__ == "__main__":
    main()
